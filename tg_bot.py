
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Telegram бот для получения Excel (.xlsx) по запросу query.

Команды:
- /start, /help — инструкции
- Любой текст — трактуется как query для поиска на WB, бот вернёт .xlsx файл

Требования:
- python-telegram-bot >= 21.4
- openpyxl >= 3.1.5
- Файл cookies.txt в корне проекта (по умолчанию)
"""

import io
import logging
import os
from datetime import datetime
from typing import List, Dict, Any

from telegram import Update, InputFile
from telegram.constants import ChatAction
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

from wb_parser import WBParser

# Логирование
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
COOKIES_FILE = os.getenv('COOKIES_FILE', 'cookies.txt')


def products_to_xlsx_bytes(products: List[Dict[str, Any]]) -> bytes:
    """Готовит Excel (.xlsx) в памяти.
    Колонки: Ссылка, Название, Количество продаж, Изображение 1..N (IMAGE("url";1))
    Высота строк ~180px, ширина колонок с изображениями ~240px.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # Определяем максимальное число изображений
    max_images = 0
    for p in products:
        imgs = p.get('image_urls', []) or []
        if len(imgs) > max_images:
            max_images = len(imgs)

    wb = Workbook()
    ws = wb.active
    ws.title = "WB"

    headers = ["Ссылка", "Название", "Количество продаж"] + [f"Изображение {i}" for i in range(1, max_images + 1)]
    ws.append(headers)

    for p in products:
        product_id = p.get('id', '')
        url = f"https://www.wildberries.ru/catalog/{product_id}/detail.aspx" if product_id else ''
        name = p.get('name', '')
        sales = p.get('sales', 0)
        images = (p.get('image_urls', []) or [])[:max_images]
        # Формируем строку без изображений
        row = [url, name, sales]
        # Добавляем ячейки с формулами IMAGE()
        for img in images:
            # Формула Excel: =IMAGE("url"; 1)
            row.append(f"=IMAGE(\"{img}\"; 1)")
        # Если изображений меньше максимума — добиваем пустыми
        if len(images) < max_images:
            row += [""] * (max_images - len(images))
        ws.append(row)

    # Устанавливаем размеры: высота строк и ширина колонок
    # Excel измеряет высоту в пунктах (~0.75 pt на пиксель при 96 DPI)
    row_height_points = 180 * 0.75  # ≈ 142.5 pt
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = row_height_points

    # Ширина колонок: приблизим 240 px к ширине Excel (~ px/7)
    image_col_width = round(240 / 7.0, 2)  # ≈ 34.29

    # Зададим разумные ширины для первых трёх колонок
    base_widths = [45, 50, 18]
    for idx, width in enumerate(base_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Колонки изображений начинаются с 4-й
    for col_idx in range(4, 4 + max_images):
        ws.column_dimensions[get_column_letter(col_idx)].width = image_col_width

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (
        "Привет! Пришлите текст запроса (query), например:\n"
        "куртка женская черная\n\n"
        "Я выполню парсинг и пришлю .xlsx файл с колонками: \n"
        "Ссылка, Название, Количество продаж, Изображение 1..N (в ячейках формула IMAGE())."
    )
    await update.message.reply_text(text)


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await start(update, context)


async def handle_query(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = (update.message.text or '').strip()
    if not query:
        await update.message.reply_text("Пустой запрос. Пришлите текст запроса.")
        return

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.UPLOAD_DOCUMENT)

    try:
        with open(COOKIES_FILE, 'r') as f:
            mayak_cookies = f.read().strip()
        if not mayak_cookies:
            await update.message.reply_text("Файл cookies пуст. Заполните cookies.txt")
            return
    except FileNotFoundError:
        await update.message.reply_text("Файл cookies.txt не найден в корне проекта")
        return
    except Exception as e:
        logger.exception("Ошибка чтения cookies")
        await update.message.reply_text(f"Ошибка чтения cookies: {e}")
        return

    try:
        parser = WBParser(mayak_cookies=mayak_cookies)
        products = parser.get_products_detailed_info_with_pics(query=query, page=1, max_products=20)
        if not products:
            await update.message.reply_text("Ничего не найдено или ошибка при получении данных.")
            return

        xlsx_bytes = products_to_xlsx_bytes(products)
        filename = f"wb_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        await update.message.reply_document(document=InputFile(io.BytesIO(xlsx_bytes), filename=filename),
                                            caption=f"Результат для запроса: {query}")
    except Exception as e:
        logger.exception("Ошибка в обработке запроса")
        await update.message.reply_text(f"Ошибка: {e}")


def build_app() -> Application:
    token = TELEGRAM_BOT_TOKEN or ""
    return Application.builder().token(token).build()


def main() -> None:
    if not TELEGRAM_BOT_TOKEN:
        logger.error("Не задан TELEGRAM_BOT_TOKEN (экспортируйте переменную окружения)")
        raise SystemExit(1)

    application = build_app()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_cmd))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_query))

    logger.info("Telegram бот запущен. Ожидание сообщений...")
    application.run_polling()


if __name__ == "__main__":
    main()
